from openpyxl import load_workbook
import shutil
from utils import get_dir, caminhos_arq
from json_parser import linhaReq2excel
#from windows import obter_numero


# Checar se linux ou NT, para detalhar caminho do diretório corretamente
def copia_arq(num_idObra):
    '''
    Função que toma o id de Obra e copia o modelo xlsx para um nome específico, retorna destino do xlsx
    '''
    caminho = get_dir()
    origem = caminho+'modelo.xlsx'
    destino = caminho+'SIEGE-ORCAMENTO-ID_'+str(num_idObra)+'.xlsx'

    shutil.copy(origem, destino)
    return destino


def insert_rows_xls(idObra, idUC, bdi, encargos):
    '''
    Função para inserir linha a linha informações de linhaReq2excel, também
    adiciona taxes (tributação) nas linhas mao de obra e materiais
    '''
    bdi = float(bdi)
    encargos = float(encargos)
    # copia modelo e renomeia para idObra
    copia_arq(idObra)
    destino = caminhos_arq(idObra)
    
    # Carregar o arquivo Excel existente
    arquivo_excel = load_workbook(destino)

    # Selecionar a planilha desejada
    iduc = f"{idUC}"
    try:
        planilha = arquivo_excel[iduc]
    except:
        planilha = arquivo_excel.create_sheet(title=iduc)

    # linha inicial do excel
    linha_inicial = 12

    # retorna tamanho original da requisição
    tamanho = linhaReq2excel(0, idObra, idUC)

    cols = len(tamanho[0])
    line = tamanho[1]


    # Logica para adicionar linhas no arquivo excel

    for row in planilha.iter_rows(min_row=linha_inicial, max_col=cols, max_row=line+linha_inicial-1):
        for cell in row:
            if cell.col_idx == 5:
                # linha de mao de obra
                res = linhaReq2excel(cell.row-linha_inicial, idObra, idUC)[0][cell.col_idx-1] 
                try:
                    cell.value = float(res)*(1+bdi/100)*(1+encargos/100)
                except:
                    cell.value = 0
                # if res is None:
                #     cell.value = 0
                # else:
                #     cell.value = float(res)*(1+bdi/100)*(1+encargos/100)
            elif cell.col_idx == 6:
                # linha de materiais
                res = linhaReq2excel(cell.row-linha_inicial, idObra, idUC)[0][cell.col_idx-1] 
                try:
                    cell.value = float(res)*(1+bdi/100)
                except:
                    cell.value = 0
                # if res is None:        
                #     cell.value = 0
                # else:
                #     cell.value = float(res)*(1+bdi/100)
            else:
                res = linhaReq2excel(cell.row-linha_inicial, idObra, idUC)[0][cell.col_idx-1]         
                cell.value = res
    # Salvar as alterações no arquivo Excel
    arquivo_excel.save(destino)

# teste para executar função que adiciona linhas no arquivo xls
# idObra = 93
# insert_rows_xls(idObra, 1)